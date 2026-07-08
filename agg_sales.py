import openpyxl, json
from collections import defaultdict

DATA_DIR = "/Users/apple/博士伦_润洁/博士伦_资料/源数据/销售数据"

# Load Taobao
wb = openpyxl.load_workbook(f'{DATA_DIR}/【淘闪-销售】Y26年H1交易情况.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]
taobao = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if not row[1]: continue
    taobao.append({
        'date': str(row[1]), 'product': str(row[3])[:50], 'city': str(row[8]),
        'sales': float(row[11] or 0), 'volume': float(row[12] or 0),
        'orders': float(row[13] or 0), 'users': float(row[15] or 0),
        'otc_rx': str(row[16]) if len(row) > 16 and row[16] else 'OTC',
        'platform': '淘闪'
    })

# Load Meituan
wb2 = openpyxl.load_workbook(f'{DATA_DIR}/【美团-销售】Y26年H1交易情况.xlsx', data_only=True)
ws2 = wb2[wb2.sheetnames[0]]
meituan = []
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
    if not row[0]: continue
    meituan.append({
        'date': str(row[0])[:6], 'product': str(row[2])[:50], 'city': str(row[5] or ''),
        'sales': float(row[6] or 0), 'orders': float(row[9] or 0),
        'users': float(row[12] or 0), 'volume': float(row[15] or 0),
        'platform': '美团', 'otc_rx': '即时零售'
    })

all_data = taobao + meituan
print(f'淘闪:{len(taobao)} 美团:{len(meituan)} 总计:{len(all_data)}')

# View 1: Platform x Month
pm = defaultdict(lambda: defaultdict(lambda: {'sales':0,'volume':0,'orders':0,'users':0}))
for d in all_data:
    m = d['date'][4:6]
    pm[d['platform']][m]['sales'] += d['sales']
    pm[d['platform']][m]['volume'] += d['volume']
    pm[d['platform']][m]['orders'] += d['orders']
    pm[d['platform']][m]['users'] += d['users']

v1 = {}
total_h1 = 0
for p in pm:
    v1[p] = {}
    pt = 0
    for m in sorted(pm[p].keys()):
        d = pm[p][m]
        v1[p][m] = {k: round(v) for k, v in d.items()}
        pt += d['sales']
    v1[p]['_total'] = round(pt)
    v1[p]['_months'] = sorted(pm[p].keys())
    total_h1 += pt
for p in v1:
    v1[p]['_share'] = round(v1[p]['_total'] / total_h1 * 100, 1)

# Calculate MoM for each platform
for p in v1:
    months = v1[p]['_months']
    prev_sales = None
    for m in months:
        cur = v1[p][m]['sales']
        if prev_sales and prev_sales > 0:
            v1[p][m]['mom'] = round((cur - prev_sales) / prev_sales * 100, 1)
        else:
            v1[p][m]['mom'] = 0
        prev_sales = cur

# View 2: Platform x OTC/RX x Month
orm = defaultdict(lambda: defaultdict(lambda: {'sales':0,'volume':0}))
for d in all_data:
    m = d['date'][4:6]
    key = d['platform'] + '|' + d['otc_rx'][:15]
    orm[key][m]['sales'] += d['sales']
    orm[key][m]['volume'] += d['volume']

v2 = {}
for key in orm:
    v2[key] = {}
    total = 0
    for m in sorted(orm[key].keys()):
        d = orm[key][m]
        v2[key][m] = {k: round(v) for k, v in d.items()}
        total += d['sales']
    v2[key]['_total'] = round(total)
    v2[key]['_months'] = sorted(orm[key].keys())

# View 4: Top 15 Cities
city_sales = defaultdict(lambda: defaultdict(float))
for d in all_data:
    if d['city']:
        city_sales[d['platform']][d['city']] += d['sales']

v4 = {}
for p in city_sales:
    top = sorted(city_sales[p].items(), key=lambda x: x[1], reverse=True)[:15]
    v4[p] = [{'city': c, 'sales': round(s)} for c, s in top]

# View 5: Product User Share
pu = defaultdict(lambda: defaultdict(float))
for d in all_data:
    pu[d['platform']][d['product'][:35]] += d.get('users', 0)

v5 = {}
for p in pu:
    total_u = sum(pu[p].values())
    top = sorted(pu[p].items(), key=lambda x: x[1], reverse=True)[:15]
    v5[p] = [{'product': n, 'users': round(u), 'share': round(u/total_u*100,1)} for n, u in top]

# Output
output = {
    'meta': {'total_sales': round(total_h1), 'period': '2026年1-6月(H1)'},
    'v1_platform_monthly': v1,
    'v2_otc_rx_monthly': v2,
    'v4_top15_cities': v4,
    'v5_product_users': v5,
}
with open('/Users/apple/字符串替换/sales_data.json', 'w') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f'Saved: {len(json.dumps(output)):,} chars')
t_share = v1.get('淘闪', {}).get('_share', 0)
m_share = v1.get('美团', {}).get('_share', 0)
print(f'H1 Total: {total_h1:,.0f} | 淘闪:{t_share}% 美团:{m_share}%')
